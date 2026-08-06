from datetime import datetime, date, timedelta
import json
import io
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, Response
from sqlalchemy import or_
from app import db
from app.models.schedule import ScheduleEntry
from app.models.request import RequestMessage
from app.models.user import User
from app.models.notification import Notification
from app.models.agenda_mom import Agenda, MOM
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


dashboard_bp = Blueprint('dashboard', __name__)


def login_required(view):
    def wrapped(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in first.', 'warning')
            return redirect(url_for('auth.login'))
        return view(*args, **kwargs)
    wrapped.__name__ = view.__name__
    return wrapped


@dashboard_bp.route('/dashboard')
@login_required
def schedule():
    today = date.today()
    selected_date = request.args.get('date')
    range_filter = request.args.get('range', 'today')
    rescheduled_only = request.args.get('rescheduled') == '1'
    current_user = User.query.get(session['user_id'])
    
    view_type = request.args.get('view_type', 'ceo')
    employee_id = request.args.get('employee_id', type=int)
    target_employee = None

    focus_date = None
    if selected_date:
        try:
            focus_date = date.fromisoformat(selected_date)
        except ValueError:
            focus_date = None

    if focus_date:
        range_filter = 'day'
        range_label = f"Schedule for {focus_date.strftime('%A, %d %b %Y')}"
        query = ScheduleEntry.query.filter(ScheduleEntry.event_date == focus_date)
    else:
        if range_filter == '7d':
            start_date = today - timedelta(days=7)
            range_label = 'Last 7 days'
        elif range_filter == '15d':
            start_date = today - timedelta(days=15)
            range_label = 'Last 15 days'
        elif range_filter == '1m':
            start_date = today - timedelta(days=30)
            range_label = 'Last month'
        elif range_filter == 'upcoming':
            start_date = today + timedelta(days=1)
            range_label = 'Upcoming'
        else:
            range_filter = 'today'
            start_date = today
            range_label = "Today's schedule"

        query = ScheduleEntry.query
        if range_filter == 'upcoming':
            query = query.filter(ScheduleEntry.event_date >= start_date)
        else:
            query = query.filter(ScheduleEntry.event_date >= start_date, ScheduleEntry.event_date <= today)

    # Filter based on view_type and user role
    if view_type == 'mine':
        query = query.filter(ScheduleEntry.user_id == current_user.id)
    elif view_type == 'employee' and current_user.is_admin_role():
        if employee_id:
            target_employee = User.query.get(employee_id)
            query = query.filter(ScheduleEntry.user_id == employee_id)
        else:
            query = query.filter(ScheduleEntry.user_id == -1) # empty
    else:
        view_type = 'ceo'
        query = query.filter(ScheduleEntry.user_id.is_(None))

    if rescheduled_only:
        query = query.filter(ScheduleEntry.reschedule == 'YES')
        range_label = f"{range_label} - Rescheduled"

    entries = query.order_by(ScheduleEntry.event_date.asc()).all()
    
    def get_start_time_sort_key(entry):
        time_str = entry.time or ''
        start_part = time_str.split('-')[0].strip() if '-' in time_str else time_str.strip()
        if not start_part:
            return (2, 0, 0)
        try:
            parsed = datetime.strptime(start_part, "%I:%M %p")
            return (0, parsed.hour, parsed.minute)
        except ValueError:
            try:
                parsed = datetime.strptime(start_part, "%H:%M")
                return (0, parsed.hour, parsed.minute)
            except ValueError:
                return (1, start_part, 0)

    entries.sort(key=lambda x: (x.event_date, get_start_time_sort_key(x), x.id))
    users = User.query.order_by(User.name.asc()).all()

    # Determine the date of the schedule in words and for the filename
    if focus_date:
        formatted_schedule_date = focus_date.strftime('%A, %d %B %Y')
        filename_date = focus_date.strftime('%Y-%m-%d')
    else:
        if range_filter == 'today':
            formatted_schedule_date = today.strftime('%A, %d %B %Y')
            filename_date = today.strftime('%Y-%m-%d')
        elif range_filter == '7d':
            formatted_schedule_date = 'Last 7 days'
            filename_date = 'last_7_days'
        elif range_filter == '15d':
            formatted_schedule_date = 'Last 15 days'
            filename_date = 'last_15_days'
        elif range_filter == '1m':
            formatted_schedule_date = 'Last month'
            filename_date = 'last_month'
        elif range_filter == 'upcoming':
            formatted_schedule_date = 'Upcoming'
            filename_date = 'upcoming'
        else:
            formatted_schedule_date = "Today's schedule"
            filename_date = today.strftime('%Y-%m-%d')

    if rescheduled_only:
        formatted_schedule_date = f"{formatted_schedule_date} (Rescheduled)"
        filename_date = f"{filename_date}_rescheduled"

    return render_template(
        'dashboard.html',
        entries=entries,
        today=today,
        current_user=current_user,
        range_filter=range_filter,
        range_label=range_label,
        rescheduled=rescheduled_only,
        view_type=view_type,
        target_employee=target_employee,
        users=users,
        formatted_schedule_date=formatted_schedule_date,
        filename_date=filename_date,
    )


@dashboard_bp.route('/generate-request', methods=['GET', 'POST'])
@login_required
def generate_request():
    if request.method == 'POST':
        subject = request.form['subject']
        priority = request.form['priority']
        discussion_date = request.form.get('discussion_date', '').strip()
        discussion_hour = request.form.get('discussion_hour', '').strip()
        discussion_minute = request.form.get('discussion_minute', '').strip()
        discussion_meridiem = request.form.get('discussion_meridiem', '').strip()
        time_needed = f"{discussion_date} {discussion_hour}:{discussion_minute} {discussion_meridiem}" if discussion_date else ''
        message = request.form.get('message', '')
        req = RequestMessage(user_id=session['user_id'], subject=subject, priority=priority, time_needed=time_needed, message=message)
        db.session.add(req)
        db.session.commit()
        flash('Request sent successfully.', 'success')
        return redirect(url_for('dashboard.request_history'))
    return render_template('generate_request.html')


@dashboard_bp.route('/request/<int:req_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_request(req_id):
    current_user = User.query.get(session['user_id'])
    req = RequestMessage.query.get_or_404(req_id)
    # Only the owner (employee) or admin (PA/SA) may edit; employees cannot edit after feedback exists
    if request.method == 'POST':
        if not (current_user.id == req.user_id or current_user.is_admin_role()):
            flash('Not authorized to edit this request.', 'danger')
            return redirect(url_for('dashboard.request_history'))
        if req.feedback:
            flash('Cannot edit a request that already has feedback.', 'warning')
            return redirect(url_for('dashboard.request_history'))

        req.subject = request.form['subject']
        req.priority = request.form['priority']
        discussion_date = request.form.get('discussion_date', '').strip()
        discussion_hour = request.form.get('discussion_hour', '').strip()
        discussion_minute = request.form.get('discussion_minute', '').strip()
        discussion_meridiem = request.form.get('discussion_meridiem', '').strip()
        req.time_needed = f"{discussion_date} {discussion_hour}:{discussion_minute} {discussion_meridiem}" if discussion_date else ''
        req.message = request.form.get('message', '')
        db.session.commit()
        flash('Request updated.', 'success')
        return redirect(url_for('dashboard.request_history'))

    discussion_date = ''
    discussion_hour = ''
    discussion_minute = ''
    discussion_meridiem = ''
    if req.time_needed:
        parts = req.time_needed.split(' ')
        if len(parts) >= 3:
            discussion_date = parts[0]
            time_part = parts[1]
            discussion_meridiem = parts[2]
            time_values = time_part.split(':')
            discussion_hour = time_values[0]
            discussion_minute = time_values[1]
    return render_template('edit_request.html', req=req, current_user=current_user, discussion_date=discussion_date, discussion_hour=discussion_hour, discussion_minute=discussion_minute, discussion_meridiem=discussion_meridiem)



@dashboard_bp.route('/request/<int:req_id>/reply', methods=['GET', 'POST'])
@login_required
def reply_request(req_id):
    current_user = User.query.get(session['user_id'])
    if not current_user.is_admin_role():
        flash('Only PA/SA can send replies.', 'danger')
        return redirect(url_for('dashboard.request_history'))

    req = RequestMessage.query.get_or_404(req_id)
    if request.method == 'POST':
        reply_date = request.form.get('reply_date', '').strip()
        reply_hour = request.form.get('reply_hour', '').strip()
        reply_minute = request.form.get('reply_minute', '').strip()
        reply_meridiem = request.form.get('reply_meridiem', '').strip()
        time_given = f"{reply_date} {reply_hour}:{reply_minute} {reply_meridiem}" if reply_date else ''
        message = request.form.get('message', '').strip()
        req.feedback_time = time_given
        req.feedback_message = message
        req.status = 'Replied'
        db.session.commit()
        flash('Reply sent to employee.', 'success')
        return redirect(url_for('dashboard.request_history'))

    return render_template('reply_request.html', req=req, current_user=current_user)


@dashboard_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    current_user = User.query.get(session['user_id'])
    if request.method == 'POST':
        name = request.form['name'].strip()
        department = request.form['department'].strip()
        designation = request.form['designation'].strip()
        email = request.form['email'].strip()
        password = request.form.get('password', '').strip()

        # Check if email is already taken by another user
        existing_user = User.query.filter(User.email == email, User.id != current_user.id).first()
        if existing_user:
            flash('This email is already in use by another account.', 'danger')
            return redirect(url_for('dashboard.profile'))

        current_user.name = name
        current_user.department = department
        current_user.designation = designation
        current_user.email = email
        if password:
            current_user.password = password
            
        db.session.commit()
        session['user_name'] = current_user.name
        flash('Profile updated successfully.', 'success')
        return redirect(url_for('dashboard.profile'))

    return render_template('profile.html', current_user=current_user)


@dashboard_bp.route('/request-history')
@login_required
def request_history():
    current_user = User.query.get(session['user_id'])
    if current_user.is_admin_role():
        requests = RequestMessage.query.order_by(RequestMessage.created_at.desc()).all()
    else:
        requests = RequestMessage.query.filter_by(user_id=session['user_id']).order_by(RequestMessage.created_at.desc()).all()
    return render_template('request_history.html', requests=requests, current_user=current_user)


@dashboard_bp.route('/feedback-history', methods=['GET', 'POST'])
@login_required
def feedback_history():
    current_user = User.query.get(session['user_id'])
    if request.method == 'POST':
        request_id = request.form.get('request_id')
        req = RequestMessage.query.get(request_id)
        if not req:
            flash('Request not found.', 'danger')
            return redirect(url_for('dashboard.feedback_history'))

        if current_user.is_admin_role():
            feedback = request.form['feedback']
            status = request.form['status']
            req.feedback = feedback
            req.status = status
            db.session.commit()
            flash('Feedback updated.', 'success')
        elif request.form.get('action') == 'acknowledge':
            req.status = 'Acknowledged'
            db.session.commit()
            flash('Acknowledged successfully.', 'success')

    feedback_query = RequestMessage.query.filter(
        or_(
            RequestMessage.feedback.isnot(None),
            RequestMessage.feedback != '',
            RequestMessage.feedback_message.isnot(None),
            RequestMessage.feedback_message != '',
            RequestMessage.feedback_time.isnot(None),
            RequestMessage.feedback_time != ''
        )
    )

    if current_user.is_admin_role():
        requests = feedback_query.order_by(RequestMessage.created_at.desc()).all()
    else:
        requests = feedback_query.filter(RequestMessage.user_id == current_user.id).order_by(RequestMessage.created_at.desc()).all()
    return render_template('feedback_history.html', requests=requests, current_user=current_user)


@dashboard_bp.route('/calendar')
@login_required
def calendar_view():
    current_user = User.query.get(session['user_id'])
    if current_user.is_admin_role():
        # SA/PA can see CEO schedule on the calendar
        entries = ScheduleEntry.query.filter(ScheduleEntry.user_id.is_(None)).order_by(ScheduleEntry.event_date.asc()).all()
    else:
        # Employees see CEO schedule and their own personal schedule
        entries = ScheduleEntry.query.filter(
            or_(
                ScheduleEntry.user_id.is_(None),
                ScheduleEntry.user_id == current_user.id
            )
        ).order_by(ScheduleEntry.event_date.asc()).all()

    events = [
        {
            'date': entry.event_date.strftime('%Y-%m-%d'),
            'activity': entry.activity,
            'priority': entry.priority
        }
        for entry in entries
    ]
    return render_template('calendar.html', events=events)


@dashboard_bp.route('/employees')
@login_required
def employees_list():
    current_user = User.query.get(session['user_id'])
    if not current_user.is_admin_role():
        flash('Only Personal Assistant or Senior Assistant can view employees.', 'danger')
        return redirect(url_for('dashboard.schedule'))
    employees = User.query.filter_by(role='Employee').order_by(User.name.asc()).all()
    return render_template('employees.html', employees=employees, current_user=current_user)


@dashboard_bp.route('/schedule/new', methods=['GET', 'POST'])
@login_required
def new_employee_schedule():
    current_user = User.query.get(session['user_id'])
    if request.method == 'POST':
        event_date_str = request.form['event_date']
        try:
            event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
        except ValueError:
            event_date = date.fromisoformat(event_date_str)

        start_time = request.form.get('start_hour', '').strip()
        start_minute = request.form.get('start_minute', '').strip()
        start_meridiem = request.form.get('start_meridiem', '').strip()
        end_time = request.form.get('end_hour', '').strip()
        end_minute = request.form.get('end_minute', '').strip()
        end_meridiem = request.form.get('end_meridiem', '').strip()

        start_value = f"{start_time}:{start_minute} {start_meridiem}" if start_time and start_minute else ''
        end_value = f"{end_time}:{end_minute} {end_meridiem}" if end_time and end_minute else ''
        time_value = f"{start_value} - {end_value}" if start_value and end_value else (start_value or end_value or '')

        activity = request.form.get('activity', 'Personal Duty')
        location = request.form.get('location', 'Office')
        google_location = request.form.get('google_location', '').strip()
        priority = request.form.get('priority', 'Normal')
        status = request.form.get('status', 'Planned')
        remark = request.form.get('remark', '')
        reschedule = request.form.get('reschedule', 'NO')
        given_time = request.form.get('given_time', '').strip()

        entry = ScheduleEntry(
            event_date=event_date,
            time=time_value,
            activity=activity,
            location=location,
            google_location=google_location,
            responsible_person=current_user.name,
            priority=priority,
            status=status,
            remark=remark,
            reschedule=reschedule,
            given_time=given_time,
            user_id=current_user.id
        )
        db.session.add(entry)
        db.session.commit()
        flash('Schedule entry created successfully.', 'success')
        return redirect(url_for('dashboard.schedule', view_type='mine', date=event_date.strftime('%Y-%m-%d')))

    default_date = request.args.get('date')
    if not default_date:
        default_date = date.today().strftime('%Y-%m-%d')
    return render_template('employee_schedule_form.html', entry=None, default_date=default_date)


@dashboard_bp.route('/schedule/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee_schedule(entry_id):
    current_user = User.query.get(session['user_id'])
    entry = ScheduleEntry.query.get_or_404(entry_id)

    # Must be owner of the schedule or SA/PA
    if not (entry.user_id == current_user.id or current_user.is_admin_role()):
        flash('Not authorized to edit this schedule entry.', 'danger')
        return redirect(url_for('dashboard.schedule'))

    if request.method == 'POST':
        event_date_str = request.form['event_date']
        try:
            entry.event_date = datetime.strptime(event_date_str, '%Y-%m-%d').date()
        except ValueError:
            entry.event_date = date.fromisoformat(event_date_str)

        start_time = request.form.get('start_hour', '').strip()
        start_minute = request.form.get('start_minute', '').strip()
        start_meridiem = request.form.get('start_meridiem', '').strip()
        end_time = request.form.get('end_hour', '').strip()
        end_minute = request.form.get('end_minute', '').strip()
        end_meridiem = request.form.get('end_meridiem', '').strip()

        start_value = f"{start_time}:{start_minute} {start_meridiem}" if start_time and start_minute else ''
        end_value = f"{end_time}:{end_minute} {end_meridiem}" if end_time and end_minute else ''
        entry.time = f"{start_value} - {end_value}" if start_value and end_value else (start_value or end_value or entry.time)

        entry.activity = request.form.get('activity', entry.activity)
        entry.location = request.form.get('location', entry.location)
        entry.google_location = request.form.get('google_location', '').strip()
        entry.priority = request.form.get('priority', entry.priority)
        entry.status = request.form.get('status', entry.status)
        entry.remark = request.form.get('remark', entry.remark)
        entry.reschedule = request.form.get('reschedule', entry.reschedule)
        entry.given_time = request.form.get('given_time', '').strip()

        db.session.commit()
        flash('Schedule entry updated.', 'success')

        if entry.user_id == current_user.id:
            return redirect(url_for('dashboard.schedule', view_type='mine'))
        else:
            return redirect(url_for('dashboard.schedule', view_type='employee', employee_id=entry.user_id))

    # Parse times for editing
    start_time, end_time = '', ''
    if entry.time:
        parts = [part.strip() for part in entry.time.split('-')]
        if len(parts) >= 2:
            start_time, end_time = parts[0], parts[1]
        else:
            start_time = parts[0]

    start_hour, start_minute, start_meridiem = '', '', ''
    if start_time and ':' in start_time:
        try:
            time_part, meridiem_part = start_time.split()
            h, m = time_part.split(':')
            start_hour = h
            start_minute = m
            start_meridiem = meridiem_part
        except Exception:
            pass

    end_hour, end_minute, end_meridiem = '', '', ''
    if end_time and ':' in end_time:
        try:
            time_part, meridiem_part = end_time.split()
            h, m = time_part.split(':')
            end_hour = h
            end_minute = m
            end_meridiem = meridiem_part
        except Exception:
            pass

    return render_template(
        'employee_schedule_form.html',
        entry=entry,
        start_hour=start_hour,
        start_minute=start_minute,
        start_meridiem=start_meridiem,
        end_hour=end_hour,
        end_minute=end_minute,
        end_meridiem=end_meridiem
    )


@dashboard_bp.route('/schedule/<int:entry_id>/delete')
@login_required
def delete_employee_schedule(entry_id):
    current_user = User.query.get(session['user_id'])
    entry = ScheduleEntry.query.get_or_404(entry_id)

    # Must be owner of the schedule or SA/PA
    if not (entry.user_id == current_user.id or current_user.is_admin_role()):
        flash('Not authorized to delete this schedule entry.', 'danger')
        return redirect(url_for('dashboard.schedule'))

    db.session.delete(entry)
    db.session.commit()
    flash('Schedule entry deleted.', 'success')

    if entry.user_id == current_user.id:
        return redirect(url_for('dashboard.schedule', view_type='mine'))
    else:
        return redirect(url_for('dashboard.schedule', view_type='employee', employee_id=entry.user_id))


@dashboard_bp.route('/employees/<int:employee_id>/notify', methods=['GET', 'POST'])
@login_required
def notify_employee(employee_id):
    current_user = User.query.get(session['user_id'])
    if not current_user.is_admin_role():
        flash('Only Personal Assistant or Senior Assistant can notify employees.', 'danger')
        return redirect(url_for('dashboard.schedule'))

    employee = User.query.get_or_404(employee_id)
    if employee.role != 'Employee':
        flash('Can only notify employees.', 'warning')
        return redirect(url_for('dashboard.employees_list'))

    # Fetch CEO schedule entries to allow linking an invitation
    ceo_schedules = ScheduleEntry.query.filter(ScheduleEntry.user_id.is_(None)).order_by(ScheduleEntry.event_date.desc(), ScheduleEntry.id.desc()).all()

    if request.method == 'POST':
        message = request.form.get('message', '').strip()
        notification_date = request.form.get('notification_date', '').strip()
        notification_hour = request.form.get('notification_hour', '').strip()
        notification_minute = request.form.get('notification_minute', '').strip()
        notification_meridiem = request.form.get('notification_meridiem', '').strip()
        schedule_entry_id = request.form.get('schedule_entry_id', '').strip()

        time_value = f"{notification_date} {notification_hour}:{notification_minute} {notification_meridiem}" if notification_date else ''

        if not message or not time_value:
            flash('All fields are required.', 'danger')
            return redirect(url_for('dashboard.notify_employee', employee_id=employee_id))

        notif = Notification(
            user_id=employee.id,
            sender_id=current_user.id,
            message=message,
            time_needed=time_value,
            acknowledged=False
        )
        if schedule_entry_id:
            notif.schedule_entry_id = int(schedule_entry_id)

        db.session.add(notif)
        db.session.commit()
        flash(f'Notification sent to {employee.name} successfully.', 'success')
        return redirect(url_for('dashboard.employees_list'))

    return render_template('notify_employee.html', employee=employee, ceo_schedules=ceo_schedules)


@dashboard_bp.route('/notification/<int:notif_id>/acknowledge', methods=['POST'])
@login_required
def acknowledge_notification(notif_id):
    notif = Notification.query.get_or_404(notif_id)
    if notif.user_id != session['user_id']:
        flash('Not authorized to acknowledge this notification.', 'danger')
        return redirect(url_for('dashboard.schedule'))

    notif.acknowledged = True

    # Automatically add to attendants list if this notification is associated with a schedule entry (meeting)
    if notif.schedule_entry_id:
        entry = ScheduleEntry.query.get(notif.schedule_entry_id)
        if entry:
            user = User.query.get(notif.user_id)
            if user and user not in entry.attendants:
                entry.attendants.append(user)

    db.session.commit()
    flash('Notification acknowledged.', 'success')
    return redirect(url_for('dashboard.notifications'))


@dashboard_bp.route('/notifications')
@login_required
def notifications():
    current_user = User.query.get(session['user_id'])
    if current_user.is_admin_role():
        acknowledged_notifs = Notification.query.filter_by(
            acknowledged=True
        ).order_by(Notification.created_at.desc()).all()
    else:
        acknowledged_notifs = Notification.query.filter_by(
            user_id=current_user.id,
            acknowledged=True
        ).order_by(Notification.created_at.desc()).all()
    
    return render_template('notifications.html', notifications=acknowledged_notifs, current_user=current_user)


@dashboard_bp.route('/schedule/<int:entry_id>/toggle-attendance', methods=['POST'])
@login_required
def toggle_attendance(entry_id):
    from flask import jsonify
    entry = ScheduleEntry.query.get_or_404(entry_id)
    current_user = User.query.get(session['user_id'])
    
    if current_user in entry.attendants:
        entry.attendants.remove(current_user)
        action = 'removed'
    else:
        entry.attendants.append(current_user)
        action = 'added'
        
    db.session.commit()
    return jsonify({
        'status': 'success',
        'action': action,
        'user_name': current_user.name,
        'user_id': current_user.id,
        'attendants': [u.name for u in entry.attendants]
    })


@dashboard_bp.route('/schedule/<int:entry_id>/toggle-user-attendance/<int:user_id>', methods=['POST'])
@login_required
def toggle_user_attendance(entry_id, user_id):
    from flask import jsonify
    current_user = User.query.get(session['user_id'])
    if not current_user.is_admin_role():
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
        
    entry = ScheduleEntry.query.get_or_404(entry_id)
    target_user = User.query.get_or_404(user_id)
    
    if target_user in entry.attendants:
        entry.attendants.remove(target_user)
        action = 'removed'
    else:
        entry.attendants.append(target_user)
        action = 'added'
        
    db.session.commit()
    return jsonify({
        'status': 'success',
        'action': action,
        'user_name': target_user.name,
        'user_id': target_user.id,
        'attendants': [u.name for u in entry.attendants]
    })


# -------------------------------------------------------------
# AGENDA & MOM MODULES
# -------------------------------------------------------------

import re

def get_safe_filename(prefix, event_date, activity_name, extension):
    if hasattr(event_date, 'strftime'):
        date_str = event_date.strftime('%Y-%m-%d')
    else:
        date_str = str(event_date).strip()
    
    clean_activity = re.sub(r'[^a-zA-Z0-9_\-]', '_', activity_name)
    clean_activity = re.sub(r'_+', '_', clean_activity).strip('_')
    if not clean_activity:
        clean_activity = "activity"
        
    return f"{prefix}_{date_str}_{clean_activity}.{extension}"


def get_download_filename(prefix, date_val, extension):
    if not date_val:
        date_str = "no_date"
    elif hasattr(date_val, 'strftime'):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = str(date_val).strip()
        date_str = re.sub(r'[^a-zA-Z0-9_\-]', '_', date_str)
        date_str = re.sub(r'_+', '_', date_str).strip('_')
        if not date_str:
            date_str = "date"
    return f"{prefix.lower()}_{date_str}.{extension}"


def generate_mom_pdf_data(mom):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#123a6b'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'DocSection',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#123a6b'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#15253d')
    )

    story.append(Paragraph("Minutes of Meeting (MOM)", title_style))
    story.append(Spacer(1, 10))
    
    chairperson_val = mom.schedule_entry.responsible_person if (mom.schedule_entry and mom.schedule_entry.responsible_person) else "-"
    attendants_list = []
    if mom.schedule_entry and mom.schedule_entry.attendants:
        attendants_list = [u.name for u in mom.schedule_entry.attendants]
    attendants_val = ", ".join(attendants_list) if attendants_list else "-"
    
    meta_data = [
        [Paragraph("<b>Meeting Date:</b>", body_style), Paragraph(format_text_for_pdf(mom.meeting_date or "-"), body_style)],
        [Paragraph("<b>Time:</b>", body_style), Paragraph(format_text_for_pdf(mom.time or "-"), body_style)],
        [Paragraph("<b>Location:</b>", body_style), Paragraph(format_text_for_pdf(mom.location or "-"), body_style)],
        [Paragraph("<b>Chairperson:</b>", body_style), Paragraph(format_text_for_pdf(chairperson_val), body_style)],
        [Paragraph("<b>Attendants:</b>", body_style), Paragraph(format_text_for_pdf(attendants_val), body_style)],
        [Paragraph("<b>Meeting Agenda:</b>", body_style), Paragraph(format_text_for_pdf(mom.meeting_agenda or "-"), body_style)]
    ]
    
    meta_table = Table(meta_data, colWidths=[100, 420])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#dce6f1')),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("Discussion Details", section_style))
    
    mom_data = [[
        Paragraph("<b>Sr.No</b>", body_style),
        Paragraph("<b>Topic</b>", body_style),
        Paragraph("<b>Points Discussed</b>", body_style),
        Paragraph("<b>Actionable Items</b>", body_style),
        Paragraph("<b>Responsibility</b>", body_style)
    ]]
    
    rows = []
    if mom.mom_table:
        try:
            rows = json.loads(mom.mom_table)
        except Exception:
            pass
            
    if not rows:
        mom_data.append([Paragraph("-", body_style), Paragraph("-", body_style), Paragraph("-", body_style), Paragraph("-", body_style), Paragraph("-", body_style)])
    else:
        for r in rows:
            sr_no = r.get('sr_no', '')
            if sr_no and "Agenda" not in str(sr_no):
                sr_no = f"Agenda {sr_no}"
            mom_data.append([
                Paragraph(format_text_for_pdf(sr_no), body_style),
                Paragraph(format_text_for_pdf(r.get('topic', '')), body_style),
                Paragraph(format_text_for_pdf(r.get('points', '')), body_style),
                Paragraph(format_text_for_pdf(r.get('actionable', '')), body_style),
                Paragraph(format_text_for_pdf(r.get('responsibility', '')), body_style)
            ])
            
    mom_table_obj = Table(mom_data, colWidths=[65, 110, 165, 110, 70])
    mom_table_obj.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f3f6fb')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dce6f1')),
    ]))
    story.append(mom_table_obj)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_agenda_excel_data(agenda):
    wb = Workbook()
    ws = wb.active
    ws.title = "Meeting Agenda"
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='123A6B')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Segoe UI', size=11, bold=True)
    regular_font = Font(name='Segoe UI', size=11)
    
    header_fill = PatternFill(start_color='123A6B', end_color='123A6B', fill_type='solid')
    
    thin_side = Side(border_style="thin", color="DCE6F1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    ws.merge_cells('A1:C1')
    ws['A1'] = agenda.title or "MEETING AGENDA"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    meta_rows = [
        ("Date:", agenda.event_date or "-"),
        ("Time:", agenda.time or "-"),
        ("Location:", agenda.location or "-"),
        ("Chairperson:", agenda.chairperson or "-"),
        ("Attendants:", agenda.attendants or "-"),
        ("Objective of Meeting:", agenda.objective or "-")
    ]
    
    curr_row = 3
    for label, val in meta_rows:
        ws.cell(row=curr_row, column=1, value=label).font = bold_font
        ws.cell(row=curr_row, column=2, value=val).font = regular_font
        ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
        curr_row += 1
        
    curr_row += 1
    
    ws.cell(row=curr_row, column=1, value="2. Schedule of Activity / Meeting").font = bold_font
    curr_row += 1
    
    headers = ["Time", "Activity / Topic", "Presenter / Responsible"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_all
        
    ws.row_dimensions[curr_row].height = 25
    
    rows = []
    if agenda.schedule_table:
        try:
            rows = json.loads(agenda.schedule_table)
        except Exception:
            pass
            
    for r in rows:
        curr_row += 1
        ws.cell(row=curr_row, column=1, value=r.get('time', ''))
        ws.cell(row=curr_row, column=2, value=r.get('activity', ''))
        ws.cell(row=curr_row, column=3, value=r.get('presenter', ''))
        
        for col_idx in range(1, 4):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = border_all
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
    curr_row += 2
    
    ws.cell(row=curr_row, column=1, value="3. Follow-Ups of Previous MOM & Tasks").font = bold_font
    ws.cell(row=curr_row, column=2, value=agenda.follow_ups or "None").font = regular_font
    ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
    
    curr_row += 2
    
    ws.cell(row=curr_row, column=1, value="4. Tasks to be Done").font = bold_font
    ws.cell(row=curr_row, column=2, value=agenda.tasks or "None").font = regular_font
    ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def format_text_for_pdf(text):
    if not text:
        return ""
    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return text.replace('\n', '<br/>')


def format_text_for_word(text):
    if not text:
        return ""
    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return text.replace('\n', '<br/>')


def make_agenda_word_html(agenda):
    rows = []
    if agenda.schedule_table:
        try:
            rows = json.loads(agenda.schedule_table)
        except Exception:
            pass
            
    table_rows_html = ""
    for r in rows:
        table_rows_html += f"""
        <tr>
            <td style="width: 25%;">{format_text_for_word(r.get('time', ''))}</td>
            <td style="width: 50%;">{format_text_for_word(r.get('activity', ''))}</td>
            <td style="width: 25%;">{format_text_for_word(r.get('presenter', ''))}</td>
        </tr>
        """
        
    html = f"""
    <html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:w="urn:schemas-microsoft-com:office:word" xmlns="http://www.w3.org/TR/REC-html40">
    <head>
    <meta charset="utf-8">
    <title>{agenda.title or "Meeting Agenda"}</title>
    <!--[if gte mso 9]>
    <xml>
      <w:WordDocument>
        <w:View>Print</w:View>
        <w:Zoom>100</w:Zoom>
      </w:WordDocument>
    </xml>
    <![endif]-->
    <style>
      body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #15253d; margin: 1in; }}
      h1 {{ color: #123a6b; font-size: 20pt; text-align: center; margin-bottom: 20pt; }}
      h2 {{ color: #123a6b; font-size: 14pt; border-bottom: 2px solid #123a6b; padding-bottom: 3pt; margin-top: 20pt; margin-bottom: 10pt; }}
      table {{ width: 100%; border-collapse: collapse; margin-top: 10pt; margin-bottom: 10pt; }}
      th, td {{ border: 1px solid #dce6f1; padding: 8pt; text-align: left; vertical-align: top; font-size: 10pt; }}
      th {{ background-color: #f3f6fb; font-weight: bold; color: #123a6b; }}
      .meta-table {{ margin-bottom: 15pt; }}
      .meta-table td {{ border: none; padding: 4pt 0; }}
      .meta-label {{ font-weight: bold; width: 120pt; color: #5f6f86; }}
      .text-block {{ line-height: 1.5; font-size: 10pt; }}
    </style>
    </head>
    <body>
      <h1>{agenda.title or "Meeting Agenda"}</h1>
      
      <table class="meta-table">
        <tr><td class="meta-label">Date:</td><td>{format_text_for_word(agenda.event_date or "-")}</td></tr>
        <tr><td class="meta-label">Time:</td><td>{format_text_for_word(agenda.time or "-")}</td></tr>
        <tr><td class="meta-label">Location:</td><td>{format_text_for_word(agenda.location or "-")}</td></tr>
        <tr><td class="meta-label">Chairperson:</td><td>{format_text_for_word(agenda.chairperson or "-")}</td></tr>
        <tr><td class="meta-label">Attendants:</td><td>{format_text_for_word(agenda.attendants or "-")}</td></tr>
      </table>
      
      <h2>1. Objective of Meeting</h2>
      <div class="text-block">{format_text_for_word(agenda.objective or "No objective specified.")}</div>
      
      <h2>2. Schedule of Activity / Meeting</h2>
      <table>
        <thead>
          <tr>
            <th style="width: 25%;">Time</th>
            <th style="width: 50%;">Activity / Topic</th>
            <th style="width: 25%;">Presenter / Responsible</th>
          </tr>
        </thead>
        <tbody>
          {table_rows_html or "<tr><td colspan='3'>None</td></tr>"}
        </tbody>
      </table>
      
      <h2>3. Follow-Ups of Previous MOM & Tasks</h2>
      <div class="text-block">{format_text_for_word(agenda.follow_ups or "None")}</div>
      
      <h2>4. Tasks to be Done</h2>
      <div class="text-block">{format_text_for_word(agenda.tasks or "None")}</div>
    </body>
    </html>
    """
    return html


def make_mom_word_html(mom):
    rows = []
    if mom.mom_table:
        try:
            rows = json.loads(mom.mom_table)
        except Exception:
            pass
            
    table_rows_html = ""
    for r in rows:
        sr_no = r.get('sr_no', '')
        if sr_no and "Agenda" not in str(sr_no):
            sr_no = f"Agenda {sr_no}"
        table_rows_html += f"""
        <tr>
            <td style="width: 8%; text-align: center;">{sr_no}</td>
            <td style="width: 22%;">{format_text_for_word(r.get('topic', ''))}</td>
            <td style="width: 35%;">{format_text_for_word(r.get('points', ''))}</td>
            <td style="width: 20%;">{format_text_for_word(r.get('actionable', ''))}</td>
            <td style="width: 15%;">{format_text_for_word(r.get('responsibility', ''))}</td>
        </tr>
        """
        
    chair_val = mom.schedule_entry.responsible_person if (mom.schedule_entry and mom.schedule_entry.responsible_person) else "-"
    att_val = ", ".join([u.name for u in mom.schedule_entry.attendants]) if (mom.schedule_entry and mom.schedule_entry.attendants) else "-"
        
    html = f"""
    <html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:w="urn:schemas-microsoft-com:office:word" xmlns="http://www.w3.org/TR/REC-html40">
    <head>
    <meta charset="utf-8">
    <title>Minutes of Meeting (MOM)</title>
    <!--[if gte mso 9]>
    <xml>
      <w:WordDocument>
        <w:View>Print</w:View>
        <w:Zoom>100</w:Zoom>
      </w:WordDocument>
    </xml>
    <![endif]-->
    <style>
      body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #15253d; margin: 1in; }}
      h1 {{ color: #123a6b; font-size: 20pt; text-align: center; margin-bottom: 20pt; }}
      h2 {{ color: #123a6b; font-size: 14pt; border-bottom: 2px solid #123a6b; padding-bottom: 3pt; margin-top: 20pt; margin-bottom: 10pt; }}
      table {{ width: 100%; border-collapse: collapse; margin-top: 10pt; margin-bottom: 10pt; }}
      th, td {{ border: 1px solid #dce6f1; padding: 8pt; text-align: left; vertical-align: top; font-size: 10pt; }}
      th {{ background-color: #f3f6fb; font-weight: bold; color: #123a6b; }}
      .meta-table {{ margin-bottom: 15pt; }}
      .meta-table td {{ border: none; padding: 4pt 0; }}
      .meta-label {{ font-weight: bold; width: 120pt; color: #5f6f86; }}
      .text-block {{ line-height: 1.5; font-size: 10pt; }}
    </style>
    </head>
    <body>
      <h1>Minutes of Meeting (MOM)</h1>
      
      <table class="meta-table">
        <tr><td class="meta-label">Meeting Date:</td><td>{format_text_for_word(mom.meeting_date or "-")}</td></tr>
        <tr><td class="meta-label">Time:</td><td>{format_text_for_word(mom.time or "-")}</td></tr>
        <tr><td class="meta-label">Location:</td><td>{format_text_for_word(mom.location or "-")}</td></tr>
        <tr><td class="meta-label">Chairperson:</td><td>{format_text_for_word(chair_val)}</td></tr>
        <tr><td class="meta-label">Attendants:</td><td>{format_text_for_word(att_val)}</td></tr>
        <tr><td class="meta-label">Meeting Agenda:</td><td>{format_text_for_word(mom.meeting_agenda or "-")}</td></tr>
      </table>
      
      <h2>Discussion Details</h2>
      <table>
        <thead>
          <tr>
            <th style="width: 8%; text-align: center;">Sr.No</th>
            <th style="width: 22%;">Topic</th>
            <th style="width: 35%;">Points Discussed</th>
            <th style="width: 20%;">Actionable Items</th>
            <th style="width: 15%;">Responsibility</th>
          </tr>
        </thead>
        <tbody>
          {table_rows_html or "<tr><td colspan='5' style='text-align: center;'>None</td></tr>"}
        </tbody>
      </table>
    </body>
    </html>
    """
    return html


def generate_agenda_pdf_data(agenda):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#123a6b'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'DocSection',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#123a6b'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#15253d')
    )

    story.append(Paragraph(format_text_for_pdf(agenda.title or "Meeting Agenda"), title_style))
    story.append(Spacer(1, 10))
    
    meta_data = [
        [Paragraph("<b>Date:</b>", body_style), Paragraph(format_text_for_pdf(agenda.event_date or "-"), body_style)],
        [Paragraph("<b>Time:</b>", body_style), Paragraph(format_text_for_pdf(agenda.time or "-"), body_style)],
        [Paragraph("<b>Location:</b>", body_style), Paragraph(format_text_for_pdf(agenda.location or "-"), body_style)],
        [Paragraph("<b>Chairperson:</b>", body_style), Paragraph(format_text_for_pdf(agenda.chairperson or "-"), body_style)],
        [Paragraph("<b>Attendants:</b>", body_style), Paragraph(format_text_for_pdf(agenda.attendants or "-"), body_style)]
    ]
    
    meta_table = Table(meta_data, colWidths=[100, 420])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#dce6f1')),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("1. Objective of Meeting", section_style))
    story.append(Paragraph(format_text_for_pdf(agenda.objective or "No objective specified."), body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("2. Schedule of Activity / Meeting", section_style))
    
    sched_data = [[
        Paragraph("<b>Time</b>", body_style),
        Paragraph("<b>Activity / Topic</b>", body_style),
        Paragraph("<b>Presenter / Responsible</b>", body_style)
    ]]
    
    rows = []
    if agenda.schedule_table:
        try:
            rows = json.loads(agenda.schedule_table)
        except Exception:
            pass
            
    if not rows:
        sched_data.append([Paragraph("-", body_style), Paragraph("-", body_style), Paragraph("-", body_style)])
    else:
        for r in rows:
            sched_data.append([
                Paragraph(format_text_for_pdf(r.get('time', '')), body_style),
                Paragraph(format_text_for_pdf(r.get('activity', '')), body_style),
                Paragraph(format_text_for_pdf(r.get('presenter', '')), body_style)
            ])
            
    sched_table = Table(sched_data, colWidths=[120, 260, 140])
    sched_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f3f6fb')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dce6f1')),
    ]))
    story.append(sched_table)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("3. Follow-Ups of Previous MOM & Tasks", section_style))
    story.append(Paragraph(format_text_for_pdf(agenda.follow_ups or "None"), body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("4. Tasks to be Done", section_style))
    story.append(Paragraph(format_text_for_pdf(agenda.tasks or "None"), body_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_mom_excel_data(mom):
    wb = Workbook()
    ws = wb.active
    ws.title = "Minutes of Meeting"
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='123A6B')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Segoe UI', size=11, bold=True)
    regular_font = Font(name='Segoe UI', size=11)
    
    header_fill = PatternFill(start_color='123A6B', end_color='123A6B', fill_type='solid')
    
    thin_side = Side(border_style="thin", color="DCE6F1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "MINUTES OF MEETING"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    chair_val = mom.schedule_entry.responsible_person if (mom.schedule_entry and mom.schedule_entry.responsible_person) else "-"
    att_val = ", ".join([u.name for u in mom.schedule_entry.attendants]) if (mom.schedule_entry and mom.schedule_entry.attendants) else "-"
    
    meta_rows = [
        ("Meeting Date:", mom.meeting_date or "-"),
        ("Time:", mom.time or "-"),
        ("Location:", mom.location or "-"),
        ("Chairperson:", chair_val),
        ("Attendants:", att_val),
        ("Meeting Agenda:", mom.meeting_agenda or "-")
    ]
    
    curr_row = 3
    for label, val in meta_rows:
        ws.cell(row=curr_row, column=1, value=label).font = bold_font
        ws.cell(row=curr_row, column=2, value=val).font = regular_font
        ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
        curr_row += 1
        
    curr_row += 1
    
    headers = ["Sr.No", "Topic", "Points Discussed", "Actionable Items", "Responsibility"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_all
        
    ws.row_dimensions[curr_row].height = 25
    
    rows = []
    if mom.mom_table:
        try:
            rows = json.loads(mom.mom_table)
        except Exception:
            pass
            
    for r in rows:
        curr_row += 1
        sr_no = r.get('sr_no', '')
        if sr_no and "Agenda" not in str(sr_no):
            sr_no = f"Agenda {sr_no}"
        ws.cell(row=curr_row, column=1, value=sr_no).alignment = Alignment(horizontal='center')
        ws.cell(row=curr_row, column=2, value=r.get('topic', ''))
        ws.cell(row=curr_row, column=3, value=r.get('points', ''))
        ws.cell(row=curr_row, column=4, value=r.get('actionable', ''))
        ws.cell(row=curr_row, column=5, value=r.get('responsibility', ''))
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = border_all
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@dashboard_bp.route('/agenda/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def agenda_edit(entry_id):
    entry = ScheduleEntry.query.get_or_404(entry_id)
    agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first()
    
    if not agenda:
        # Prepopulate default values
        title = f"Agenda: {entry.activity}"
        event_date = entry.event_date.strftime('%Y-%m-%d')
        time = entry.time or ""
        location = entry.location or ""
        chairperson = entry.responsible_person or ""
        attendants = ", ".join([u.name for u in entry.attendants])
        
        objective = f"The objective of this meeting is to discuss and coordinate on the activity: '{entry.activity}'."
        if entry.remark:
            objective += f" Additional context/remark: {entry.remark}."
            
        default_sched = [{"time": entry.time or "", "activity": entry.activity or "", "presenter": entry.responsible_person or ""}]
        schedule_table = json.dumps(default_sched)
        
        agenda = Agenda(
            schedule_entry_id=entry_id,
            title=title,
            event_date=event_date,
            time=time,
            location=location,
            chairperson=chairperson,
            attendants=attendants,
            objective=objective,
            schedule_table=schedule_table,
            follow_ups="",
            tasks=""
        )
    
    if request.method == 'POST':
        agenda.title = request.form.get('title', '').strip()
        agenda.event_date = request.form.get('event_date', '').strip()
        agenda.time = request.form.get('time', '').strip()
        agenda.location = request.form.get('location', '').strip()
        agenda.chairperson = request.form.get('chairperson', '').strip()
        agenda.attendants = request.form.get('attendants', '').strip()
        agenda.objective = request.form.get('objective', '').strip()
        agenda.follow_ups = request.form.get('follow_ups', '').strip()
        agenda.tasks = request.form.get('tasks', '').strip()
        
        # Parse dynamic schedule table rows
        times = request.form.getlist('sched_time[]')
        activities = request.form.getlist('sched_activity[]')
        presenters = request.form.getlist('sched_presenter[]')
        
        sched_rows = []
        for t, a, p in zip(times, activities, presenters):
            if t.strip() or a.strip() or p.strip():
                sched_rows.append({
                    'time': t.strip(),
                    'activity': a.strip(),
                    'presenter': p.strip()
                })
        agenda.schedule_table = json.dumps(sched_rows)
        
        if not agenda.id:
            db.session.add(agenda)
            
        db.session.commit()
        flash('Agenda saved successfully.', 'success')
        return redirect(url_for('dashboard.agenda_view', entry_id=entry_id))
        
    schedule_rows = []
    if agenda.schedule_table:
        try:
            schedule_rows = json.loads(agenda.schedule_table)
        except Exception:
            pass
            
    return render_template('agenda_form.html', entry=entry, agenda=agenda, schedule_rows=schedule_rows)


@dashboard_bp.route('/agenda/<int:entry_id>/view', methods=['GET', 'POST'])
@login_required
def agenda_view(entry_id):
    entry = ScheduleEntry.query.get_or_404(entry_id)
    agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first()
    
    if not agenda:
        flash('No agenda found for this activity. Please fill it first.', 'warning')
        return redirect(url_for('dashboard.agenda_edit', entry_id=entry_id))
        
    if request.method == 'POST':
        # Direct editing from View page
        agenda.title = request.form.get('title', '').strip()
        agenda.event_date = request.form.get('event_date', '').strip()
        agenda.time = request.form.get('time', '').strip()
        agenda.location = request.form.get('location', '').strip()
        agenda.chairperson = request.form.get('chairperson', '').strip()
        agenda.attendants = request.form.get('attendants', '').strip()
        agenda.objective = request.form.get('objective', '').strip()
        agenda.follow_ups = request.form.get('follow_ups', '').strip()
        agenda.tasks = request.form.get('tasks', '').strip()
        
        times = request.form.getlist('sched_time[]')
        activities = request.form.getlist('sched_activity[]')
        presenters = request.form.getlist('sched_presenter[]')
        
        sched_rows = []
        for t, a, p in zip(times, activities, presenters):
            if t.strip() or a.strip() or p.strip():
                sched_rows.append({
                    'time': t.strip(),
                    'activity': a.strip(),
                    'presenter': p.strip()
                })
        agenda.schedule_table = json.dumps(sched_rows)
        db.session.commit()
        
        # Check download action
        download_action = request.form.get('download_action')
        if download_action == 'pdf':
            pdf_bytes = generate_agenda_pdf_data(agenda)
            filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "pdf")
            return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
        elif download_action == 'excel':
            excel_bytes = generate_agenda_excel_data(agenda)
            filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "xlsx")
            return Response(excel_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
        elif download_action == 'word':
            word_html = make_agenda_word_html(agenda)
            filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "doc")
            return Response(word_html, mimetype="application/msword", headers={"Content-Disposition": f"attachment; filename={filename}"})
            
        flash('Agenda updated successfully.', 'success')
        return redirect(url_for('dashboard.agenda_view', entry_id=entry_id))
        
    schedule_rows = []
    if agenda.schedule_table:
        try:
            schedule_rows = json.loads(agenda.schedule_table)
        except Exception:
            pass
            
    return render_template('agenda_view.html', entry=entry, agenda=agenda, schedule_rows=schedule_rows)


@dashboard_bp.route('/agenda/<int:entry_id>/pdf')
@login_required
def agenda_pdf(entry_id):
    agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = agenda.schedule_entry
    pdf_bytes = generate_agenda_pdf_data(agenda)
    filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "pdf")
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@dashboard_bp.route('/agenda/<int:entry_id>/word')
@login_required
def agenda_word(entry_id):
    agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = agenda.schedule_entry
    word_html = make_agenda_word_html(agenda)
    filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "doc")
    return Response(
        word_html,
        mimetype="application/msword",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@dashboard_bp.route('/agenda/<int:entry_id>/excel')
@login_required
def agenda_excel(entry_id):
    agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = agenda.schedule_entry
    excel_bytes = generate_agenda_excel_data(agenda)
    filename = get_download_filename("agenda", agenda.event_date or entry.event_date, "xlsx")
    return Response(
        excel_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@dashboard_bp.route('/mom/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def mom_edit(entry_id):
    entry = ScheduleEntry.query.get_or_404(entry_id)
    mom = MOM.query.filter_by(schedule_entry_id=entry_id).first()
    
    if not mom:
        # Prepopulate default values
        meeting_date = entry.event_date.strftime('%d/%m/%Y')
        time = entry.time or ""
        location = entry.location or ""
        
        # Try to load Agenda objective as meeting agenda
        agenda = Agenda.query.filter_by(schedule_entry_id=entry_id).first()
        if agenda and agenda.objective:
            meeting_agenda = agenda.objective
        else:
            meeting_agenda = f"Discussion regarding activity: {entry.activity}"
            
        default_mom = [{"sr_no": "Agenda 1", "topic": entry.activity or "", "points": "", "actionable": "", "responsibility": entry.responsible_person or ""}]
        mom_table = json.dumps(default_mom)
        
        mom = MOM(
            schedule_entry_id=entry_id,
            meeting_date=meeting_date,
            time=time,
            location=location,
            meeting_agenda=meeting_agenda,
            mom_table=mom_table
        )
        
    if request.method == 'POST':
        mom.meeting_date = request.form.get('meeting_date', '').strip()
        mom.time = request.form.get('time', '').strip()
        mom.location = request.form.get('location', '').strip()
        mom.meeting_agenda = request.form.get('meeting_agenda', '').strip()
        
        # Parse dynamic Discussion chart
        sr_nos = request.form.getlist('mom_sr_no[]')
        topics = request.form.getlist('mom_topic[]')
        points = request.form.getlist('mom_points[]')
        actionables = request.form.getlist('mom_actionable[]')
        responsibilities = request.form.getlist('mom_responsibility[]')
        
        mom_rows = []
        for s, t, p, a, r in zip(sr_nos, topics, points, actionables, responsibilities):
            if s.strip() or t.strip() or p.strip() or a.strip() or r.strip():
                mom_rows.append({
                    'sr_no': s.strip(),
                    'topic': t.strip(),
                    'points': p.strip(),
                    'actionable': a.strip(),
                    'responsibility': r.strip()
                })
        mom.mom_table = json.dumps(mom_rows)
        
        if not mom.id:
            db.session.add(mom)
            
        db.session.commit()
        flash('MOM saved successfully.', 'success')
        return redirect(url_for('dashboard.mom_view', entry_id=entry_id))
        
    mom_rows = []
    if mom.mom_table:
        try:
            mom_rows = json.loads(mom.mom_table)
        except Exception:
            pass
            
    return render_template('mom_form.html', entry=entry, mom=mom, mom_rows=mom_rows)


@dashboard_bp.route('/mom/<int:entry_id>/view', methods=['GET', 'POST'])
@login_required
def mom_view(entry_id):
    entry = ScheduleEntry.query.get_or_404(entry_id)
    mom = MOM.query.filter_by(schedule_entry_id=entry_id).first()
    
    if not mom:
        flash('No MOM found for this activity. Please fill it first.', 'warning')
        return redirect(url_for('dashboard.mom_edit', entry_id=entry_id))
        
    if request.method == 'POST':
        # Direct editing from View page
        mom.meeting_date = request.form.get('meeting_date', '').strip()
        mom.time = request.form.get('time', '').strip()
        mom.location = request.form.get('location', '').strip()
        mom.meeting_agenda = request.form.get('meeting_agenda', '').strip()
        
        sr_nos = request.form.getlist('mom_sr_no[]')
        topics = request.form.getlist('mom_topic[]')
        points = request.form.getlist('mom_points[]')
        actionables = request.form.getlist('mom_actionable[]')
        responsibilities = request.form.getlist('mom_responsibility[]')
        
        mom_rows = []
        for s, t, p, a, r in zip(sr_nos, topics, points, actionables, responsibilities):
            if s.strip() or t.strip() or p.strip() or a.strip() or r.strip():
                mom_rows.append({
                    'sr_no': s.strip(),
                    'topic': t.strip(),
                    'points': p.strip(),
                    'actionable': a.strip(),
                    'responsibility': r.strip()
                })
        mom.mom_table = json.dumps(mom_rows)
        db.session.commit()
        
        # Check download action
        download_action = request.form.get('download_action')
        if download_action == 'pdf':
            pdf_bytes = generate_mom_pdf_data(mom)
            filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "pdf")
            return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
        elif download_action == 'excel':
            excel_bytes = generate_mom_excel_data(mom)
            filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "xlsx")
            return Response(excel_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
        elif download_action == 'word':
            word_html = make_mom_word_html(mom)
            filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "doc")
            return Response(word_html, mimetype="application/msword", headers={"Content-Disposition": f"attachment; filename={filename}"})
            
        flash('MOM updated successfully.', 'success')
        return redirect(url_for('dashboard.mom_view', entry_id=entry_id))
        
    mom_rows = []
    if mom.mom_table:
        try:
            mom_rows = json.loads(mom.mom_table)
        except Exception:
            pass
            
    return render_template('mom_view.html', entry=entry, mom=mom, mom_rows=mom_rows)


@dashboard_bp.route('/mom/<int:entry_id>/word')
@login_required
def mom_word(entry_id):
    mom = MOM.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = mom.schedule_entry
    word_html = make_mom_word_html(mom)
    filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "doc")
    return Response(
        word_html,
        mimetype="application/msword",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@dashboard_bp.route('/mom/<int:entry_id>/excel')
@login_required
def mom_excel(entry_id):
    mom = MOM.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = mom.schedule_entry
    excel_bytes = generate_mom_excel_data(mom)
    filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "xlsx")
    return Response(
        excel_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@dashboard_bp.route('/mom/<int:entry_id>/pdf')
@login_required
def mom_pdf(entry_id):
    mom = MOM.query.filter_by(schedule_entry_id=entry_id).first_or_404()
    entry = mom.schedule_entry
    pdf_bytes = generate_mom_pdf_data(mom)
    filename = get_download_filename("mom", mom.meeting_date or entry.event_date, "pdf")
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


